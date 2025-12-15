import pandas as pd


from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# 1. 读取 pandas 算好的成绩表
wb = load_workbook("班级成绩_已排序.xlsx")
ws = wb["成绩排名"]

# 2. 定义美化样式（像设计成绩单模板）
header_font = Font(name="宋体", size=11, bold=True, color="FFFFFF")  # 表头字体（白色加粗）
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # 表头蓝底
fail_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 不及格标浅红
border = Border(  # 给所有单元格加细边框
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)

# 3. 应用样式（逐部分美化）
# 美化表头（第一行：姓名、语文...）
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

# 美化内容：加边框+标红不及格（<60分）
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # 从第2行（第一个学生）开始
    for cell in row:
        cell.border = border  # 所有单元格加边框
        # 若单元格是科目分数且<60，标浅红（排除姓名、总分、排名列）
        if cell.column in [2, 3, 4] and cell.value is not None and cell.value < 60:
            cell.fill = fail_fill

# 调整列宽（避免文字挤在一起）
ws.column_dimensions["A"].width = 12  # 姓名列
ws.column_dimensions["B:C:D"].width = 10  # 各科分数列
ws.column_dimensions["E:F"].width = 12  # 总分、排名列

# 4. 保存正式成绩单
wb.save("班级成绩_正式版.xlsx")

